"""Compare two Excel files cell-by-cell and report any differences.
Usage: python tools/compare_excels.py file_before.xlsx file_after.xlsx
Exits with code 0 if identical, 1 if differences found.
"""
import sys
try:
    from openpyxl import load_workbook
except Exception as e:
    print("openpyxl is required to run this script. Install with: pip install openpyxl")
    sys.exit(2)


def cell_value_str(cell):
    if cell is None:
        return ""
    if cell.value is None:
        return ""
    return str(cell.value)


def compare_files(a_path, b_path):
    wa = load_workbook(a_path, data_only=True)
    wb = load_workbook(b_path, data_only=True)
    sheets_a = wa.sheetnames
    sheets_b = wb.sheetnames
    if sheets_a != sheets_b:
        print(f"Different sheet lists:\n  {sheets_a}\n  {sheets_b}")
        return 1
    diffs = []
    for name in sheets_a:
        sa = wa[name]
        sb = wb[name]
        max_row = max(sa.max_row, sb.max_row)
        max_col = max(sa.max_column, sb.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                va = cell_value_str(sa.cell(row=r, column=c))
                vb = cell_value_str(sb.cell(row=r, column=c))
                if va != vb:
                    diffs.append((name, r, c, va, vb))
                    if len(diffs) > 50:
                        print("More than 50 differences, stopping early.")
                        break
            if len(diffs) > 50:
                break
        if len(diffs) > 50:
            break
    if not diffs:
        print("Files are identical (cell-by-cell)")
        return 0
    print(f"Found {len(diffs)} differences (showing up to 50):")
    for s, r, c, va, vb in diffs[:50]:
        print(f"Sheet={s} R={r} C={c} | before={va!r} | after={vb!r}")
    return 1


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python tools/compare_excels.py before.xlsx after.xlsx")
        sys.exit(2)
    sys.exit(compare_files(sys.argv[1], sys.argv[2]))
