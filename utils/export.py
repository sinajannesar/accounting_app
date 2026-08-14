"""خروجی Excel و PDF"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    _OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    Font = Alignment = PatternFill = None
    _OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    _REPORTLAB_AVAILABLE = True
except Exception:
    colors = None
    A4 = landscape = None
    getSampleStyleSheet = ParagraphStyle = None
    mm = None
    pdfmetrics = TTFont = None
    SimpleDocTemplate = Table = TableStyle = Paragraph = Spacer = None
    _REPORTLAB_AVAILABLE = False

from utils.num2words_fa import amount_to_words_rial


def _format_number(n):
    if n is None:
        return ""
    if isinstance(n, float):
        if n == int(n):
            return f"{int(n):,}"
        return f"{n:,.0f}"
    return f"{int(n):,}"


def export_to_excel(file_path, title, headers, rows, totals=None):
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    if totals:
        ws.append([])
        ws.append(totals)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(file_path)


def _register_persian_font():
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for fp in font_paths:
        if Path(fp).exists():
            try:
                pdfmetrics.registerFont(TTFont("PersianFont", fp))
                return "PersianFont"
            except Exception:
                continue
    return "Helvetica"


def export_to_pdf(file_path, title, headers, rows, totals=None):
    if not _REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab is not installed. Install it with: pip install reportlab")
    font_name = _register_persian_font()
    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontName=font_name,
        fontSize=14,
        alignment=1,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        alignment=2,
    )

    elements = [Paragraph(title, title_style), Spacer(1, 10)]

    table_data = [headers]
    for row in rows:
        table_data.append([str(c) if c is not None else "" for c in row])
    if totals:
        table_data.append([str(c) if c is not None else "" for c in totals])

    col_count = len(headers)
    page_width = landscape(A4)[0] - 30 * mm
    col_width = page_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ])
    )
    elements.append(table)
    doc.build(elements)


def format_row_numbers(row, numeric_indices):
    result = list(row)
    for i in numeric_indices:
        if i < len(result):
            result[i] = _format_number(result[i])
    return result


def export_receipt_pdf(file_path, entry, company_name=None):
    """چاپ رسید تک‌سندی (فیش) شامل شماره سند، تاریخ، سررسید، شرح، سطرها و مبلغ به حروف"""
    font_name = _register_persian_font()
    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReceiptTitle", parent=styles["Heading1"],
        fontName=font_name, fontSize=15, alignment=1,
    )
    sub_style = ParagraphStyle(
        "ReceiptSub", parent=styles["Normal"],
        fontName=font_name, fontSize=10, alignment=1, textColor=colors.grey,
    )
    info_style = ParagraphStyle(
        "ReceiptInfo", parent=styles["Normal"],
        fontName=font_name, fontSize=11, alignment=2,
    )
    words_style = ParagraphStyle(
        "ReceiptWords", parent=styles["Normal"],
        fontName=font_name, fontSize=10, alignment=2, textColor=colors.HexColor("#333333"),
    )

    elements = []
    if company_name:
        elements.append(Paragraph(company_name, sub_style))
    elements.append(Paragraph("رسید سند حسابداری", title_style))
    elements.append(Spacer(1, 12))

    total_debit = sum(l["debit"] for l in entry["lines"])
    total_credit = sum(l["credit"] for l in entry["lines"])
    total_amount = max(total_debit, total_credit)

    info_data = [
        [f"شماره سند: {entry['entry_number']}", f"تاریخ: {entry['entry_date']}"],
        [f"سررسید: {entry.get('due_date') or '-'}", f"شرح: {entry.get('description') or '-'}"],
    ]
    info_table = Table(info_data, colWidths=[85 * mm, 85 * mm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    headers = ["حساب", "شرح سطر", "بدهکار", "بستانکار"]
    table_data = [headers]
    for line in entry["lines"]:
        table_data.append([
            f"{line['account_code']} - {line['account_name']}",
            line.get("line_description") or "",
            _format_number(line["debit"]) if line["debit"] else "",
            _format_number(line["credit"]) if line["credit"] else "",
        ])
    table_data.append(["", "جمع", _format_number(total_debit), _format_number(total_credit)])

    col_widths = [55 * mm, 55 * mm, 35 * mm, 35 * mm]
    lines_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    lines_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EFEFEF")),
        ("FONTNAME", (0, -1), (-1, -1), font_name),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F7F7")]),
    ]))
    elements.append(lines_table)
    elements.append(Spacer(1, 14))

    words = amount_to_words_rial(total_amount)
    elements.append(Paragraph(f"مبلغ به حروف: {words}", words_style))
    elements.append(Spacer(1, 30))

    sign_data = [["امضای دریافت‌کننده", "امضای تأییدکننده"]]
    sign_table = Table(sign_data, colWidths=[85 * mm, 85 * mm])
    sign_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 20),
        ("LINEABOVE", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(sign_table)

    doc.build(elements)